"""
Report Generator for Transition Scout
Creates Excel/CSV outputs with personalized outreach templates
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Optional
import logging
from datetime import datetime
import os

from config import OUTPUT_CONFIG, OUTREACH_TEMPLATES

class ReportGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def generate_lead_report(self, qualified_companies: List[Dict], output_path: str = None) -> str:
        """Generate comprehensive lead report with outreach templates"""
        if not qualified_companies:
            self.logger.warning("No qualified companies to generate report for")
            return None
        
        # Create output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            # Save to Downloads folder by default
            downloads_path = os.path.expanduser("~/Downloads")
            output_path = os.path.join(downloads_path, f"transition_scout_leads_{timestamp}.xlsx")
        
        # Prepare data for export
        export_data = self._prepare_export_data(qualified_companies)
        
        # Create reports based on configuration
        if OUTPUT_CONFIG["output_format"] in ["excel", "both"]:
            self._create_excel_report(export_data, output_path)
        
        if OUTPUT_CONFIG["output_format"] in ["csv", "both"]:
            # Create CSV version
            csv_path = output_path.replace('.xlsx', '.csv')
            self._create_csv_report(export_data, csv_path)
            self.logger.info(f"CSV report generated: {csv_path}")
        
        self.logger.info(f"Lead report generated: {output_path}")
        return output_path
    
    def _prepare_export_data(self, qualified_companies: List[Dict]) -> Dict:
        """Prepare data for export with all necessary columns"""
        export_data = {
            'companies': [],
            'decision_makers': [],
            'outreach_templates': []
        }
        
        for company in qualified_companies:
            # Company data
            company_row = {
                'Company Name': company.get('company_name', ''),
                'Industry': company.get('industry', ''),
                'Location': company.get('location', ''),
                'Employee Count': company.get('employee_count', ''),
                'Revenue Range': company.get('revenue_range', ''),
                'Business Age (Years)': company.get('business_age', ''),
                'Website': company.get('website', ''),
                'Phone': company.get('phone', ''),
                'LinkedIn URL': company.get('linkedin_url', ''),
                'Qualification Score': company.get('qualification_score', 0),
                'Grade': company.get('qualification_details', {}).get('score_breakdown', {}).get('grade', ''),
                'Contact Score': company.get('qualification_details', {}).get('criteria_checks', {}).get('contact_information', {}).get('score', 0)
            }
            export_data['companies'].append(company_row)
            
            # Decision makers data
            decision_makers = company.get('decision_makers', [])
            for dm in decision_makers:
                dm_row = {
                    'Company Name': company.get('company_name', ''),
                    'Decision Maker Name': dm.get('name', ''),
                    'Title': dm.get('title', ''),
                    'Email': dm.get('email', ''),
                    'Phone': dm.get('phone', ''),
                    'LinkedIn URL': dm.get('linkedin_url', '')
                }
                export_data['decision_makers'].append(dm_row)
            
            # Outreach templates
            if OUTPUT_CONFIG["include_outreach_templates"]:
                template_row = self._generate_outreach_template(company)
                export_data['outreach_templates'].append(template_row)
        
        return export_data
    
    def _generate_outreach_template(self, company: Dict) -> Dict:
        """Generate personalized outreach template for a company"""
        # Get primary decision maker
        decision_makers = company.get('decision_makers', [])
        primary_contact = decision_makers[0] if decision_makers else {}
        
        # Calculate business age
        business_age = company.get('business_age', '')
        if not business_age and company.get('founded_year'):
            try:
                current_year = datetime.now().year
                business_age = current_year - int(company['founded_year'])
            except:
                business_age = 'Unknown'
        
        # Personalize templates
        email_body = OUTREACH_TEMPLATES["email_body"].format(
            first_name=primary_contact.get('name', '').split()[0] if primary_contact.get('name') else 'there',
            company_name=company.get('company_name', ''),
            business_age=business_age,
            your_name='[Your Name]'
        )
        
        linkedin_message = OUTREACH_TEMPLATES["linkedin_message"].format(
            first_name=primary_contact.get('name', '').split()[0] if primary_contact.get('name') else 'there',
            company_name=company.get('company_name', ''),
            business_age=business_age,
            your_name='[Your Name]'
        )
        
        return {
            'Company Name': company.get('company_name', ''),
            'Primary Contact': primary_contact.get('name', ''),
            'Contact Title': primary_contact.get('title', ''),
            'Email Subject': OUTREACH_TEMPLATES["email_subject"],
            'Email Body': email_body,
            'LinkedIn Message': linkedin_message,
            'Best Contact Method': self._determine_best_contact_method(company),
            'Follow-up Timeline': '3-5 business days'
        }
    
    def _determine_best_contact_method(self, company: Dict) -> str:
        """Determine the best method to contact a company"""
        decision_makers = company.get('decision_makers', [])
        hunter_emails = company.get('hunter_emails', [])
        
        if decision_makers and any(dm.get('email') for dm in decision_makers):
            return 'Email (Direct)'
        elif hunter_emails:
            return 'Email (Hunter.io)'
        elif company.get('phone'):
            return 'Phone'
        elif company.get('linkedin_url'):
            return 'LinkedIn'
        else:
            return 'Website Contact Form'
    
    def _create_excel_report(self, export_data: Dict, output_path: str):
        """Create comprehensive Excel report with multiple sheets"""
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        self._create_companies_sheet(workbook, export_data['companies'])
        self._create_decision_makers_sheet(workbook, export_data['decision_makers'])
        
        if OUTPUT_CONFIG["include_outreach_templates"]:
            self._create_outreach_sheet(workbook, export_data['outreach_templates'])
        
        # Create summary sheet
        self._create_summary_sheet(workbook, export_data)
        
        # Save workbook
        workbook.save(output_path)
    
    def _create_companies_sheet(self, workbook: openpyxl.Workbook, companies_data: List[Dict]):
        """Create companies overview sheet"""
        sheet = workbook.create_sheet("Qualified Companies")
        
        # Convert to DataFrame and write
        df = pd.DataFrame(companies_data)
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        
        # Format headers
        self._format_headers(sheet, len(df.columns))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_decision_makers_sheet(self, workbook: openpyxl.Workbook, decision_makers_data: List[Dict]):
        """Create decision makers contact sheet"""
        sheet = workbook.create_sheet("Decision Makers")
        
        df = pd.DataFrame(decision_makers_data)
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        
        self._format_headers(sheet, len(df.columns))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_outreach_sheet(self, workbook: openpyxl.Workbook, outreach_data: List[Dict]):
        """Create outreach templates sheet"""
        sheet = workbook.create_sheet("Outreach Templates")
        
        df = pd.DataFrame(outreach_data)
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        
        self._format_headers(sheet, len(df.columns))
        
        # Auto-adjust column widths for outreach templates
        for column in sheet.columns:
            column_letter = column[0].column_letter
            if column_letter in ['E', 'F', 'G']:  # Email body, LinkedIn message columns
                sheet.column_dimensions[column_letter].width = 60
            else:
                sheet.column_dimensions[column_letter].width = 20
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, export_data: Dict):
        """Create summary statistics sheet"""
        sheet = workbook.create_sheet("Summary")
        
        # Summary statistics
        summary_data = [
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Companies Found", len(export_data['companies'])],
            ["Total Decision Makers", len(export_data['decision_makers'])],
            ["Outreach Templates Generated", len(export_data['outreach_templates'])],
            ["", ""],
            ["Target Criteria", ""],
            ["Employee Range", "50-500"],
            ["Revenue Range", "$5M - $100M"],
            ["Business Age Minimum", "15+ years"],
            ["Target Industries", ", ".join(["Manufacturing", "Construction", "Professional Services", "Healthcare"])]
        ]
        
        for row in summary_data:
            sheet.append(row)
        
        # Format summary
        for row in range(1, len(summary_data) + 1):
            if row <= 4:  # Header rows
                sheet[f"A{row}"].font = Font(bold=True)
                sheet[f"A{row}"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                sheet[f"A{row}"].font = Font(color="FFFFFF", bold=True)
    
    def _format_headers(self, sheet: Worksheet, num_columns: int):
        """Format header row with styling"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, num_columns + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def _create_csv_report(self, export_data: Dict, output_path: str):
        """Create CSV report (simplified version)"""
        # Use companies data as main CSV
        df = pd.DataFrame(export_data['companies'])
        df.to_csv(output_path, index=False)
        
        self.logger.info(f"CSV report generated: {output_path}")
    
    def generate_qualification_summary(self, companies: List[Dict]) -> Dict:
        """Generate summary statistics for the qualification process"""
        if not companies:
            return {}
        
        # Calculate statistics
        total_companies = len(companies)
        qualified_companies = [c for c in companies if c.get('qualification_score', 0) >= 70]
        qualified_count = len(qualified_companies)
        
        # Score distribution
        score_ranges = {
            'A (90-100)': 0,
            'B (80-89)': 0,
            'C (70-79)': 0,
            'D (60-69)': 0,
            'F (0-59)': 0
        }
        
        for company in companies:
            grade = company.get('qualification_details', {}).get('score_breakdown', {}).get('grade', 'F')
            if grade == 'A':
                score_ranges['A (90-100)'] += 1
            elif grade == 'B':
                score_ranges['B (80-89)'] += 1
            elif grade == 'C':
                score_ranges['C (70-79)'] += 1
            elif grade == 'D':
                score_ranges['D (60-69)'] += 1
            else:
                score_ranges['F (0-59)'] += 1
        
        return {
            'total_companies_analyzed': total_companies,
            'qualified_companies': qualified_count,
            'qualification_rate': (qualified_count / total_companies * 100) if total_companies > 0 else 0,
            'score_distribution': score_ranges,
            'average_qualification_score': sum(c.get('qualification_score', 0) for c in companies) / total_companies if total_companies > 0 else 0
        }
